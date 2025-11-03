"""
Écrivains de fichiers Excel pour l'application bancaire
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
import logging
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

logger = logging.getLogger(__name__)

class ExcelWriter:
    """Écrivain de fichiers Excel avec formatage avancé"""
    
    def __init__(self, output_dir: str = "data/outputs"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Styles par défaut
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        logger.info(f"Écrivain Excel initialisé - dossier de sortie: {self.output_dir}")
    
    def write_all_reports(self, reports_data: Dict[str, Dict[str, pd.DataFrame]], 
                         scenario_config: Dict) -> Dict[str, str]:
        """Écrire tous les rapports Excel"""
        logger.info("Écriture de tous les rapports Excel")
        
        generated_files = {}
        
        try:
            # Rapports de simulation
            if 'simulation' in reports_data:
                sim_file = self.write_simulation_report(reports_data['simulation'], scenario_config)
                generated_files['simulation'] = sim_file
            
            # Rapports comptables
            if 'accounting' in reports_data:
                acc_file = self.write_accounting_report(reports_data['accounting'], scenario_config)
                generated_files['accounting'] = acc_file
            
            # Rapports de consolidation
            if 'consolidation' in reports_data:
                cons_file = self.write_consolidation_report(reports_data['consolidation'], scenario_config)
                generated_files['consolidation'] = cons_file
            
            # Rapports FINREP
            if 'finrep' in reports_data:
                finrep_file = self.write_finrep_report(reports_data['finrep'], scenario_config)
                generated_files['finrep'] = finrep_file
            
            # Rapports COREP
            if 'corep' in reports_data:
                corep_file = self.write_corep_report(reports_data['corep'], scenario_config)
                generated_files['corep'] = corep_file
            
            # Rapports de liquidité
            if 'liquidity' in reports_data:
                liq_file = self.write_liquidity_report(reports_data['liquidity'], scenario_config)
                generated_files['liquidity'] = liq_file
            
            # Rapport RUBA
            if 'ruba' in reports_data:
                ruba_file = self.write_ruba_report(reports_data['ruba'], scenario_config)
                generated_files['ruba'] = ruba_file
            
            logger.info(f"Rapports Excel générés: {list(generated_files.keys())}")
            return generated_files
            
        except Exception as e:
            logger.error(f"Erreur lors de l'écriture des rapports: {e}")
            raise
    
    def write_simulation_report(self, simulation_data: Dict[str, pd.DataFrame], 
                              config: Dict) -> str:
        """Écrire le rapport de simulation"""
        filename = f"Simulation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Feuille de résumé
            if 'summary' in simulation_data:
                simulation_data['summary'].to_excel(writer, sheet_name='Summary', index=True)
                self._format_summary_sheet(writer.book['Summary'])
            
            # Positions détaillées
            if 'positions' in simulation_data:
                positions_sample = simulation_data['positions'].head(1000)  # Limiter à 1000 lignes
                positions_sample.to_excel(writer, sheet_name='Positions_Sample', index=False)
                self._format_data_sheet(writer.book['Positions_Sample'])
            
            # Flux de trésorerie
            if 'cash_flows' in simulation_data:
                cf_sample = simulation_data['cash_flows'].head(1000)
                cf_sample.to_excel(writer, sheet_name='CashFlows_Sample', index=False)
                self._format_data_sheet(writer.book['CashFlows_Sample'])
            
            # Dérivés
            if 'derivatives' in simulation_data:
                simulation_data['derivatives'].to_excel(writer, sheet_name='Derivatives', index=False)
                self._format_data_sheet(writer.book['Derivatives'])
            
            # Méthodologie
            methodology = self._create_simulation_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport de simulation généré: {filepath}")
        return str(filepath)
    
    def write_accounting_report(self, accounting_data: Dict[str, pd.DataFrame], 
                              config: Dict) -> str:
        """Écrire le rapport comptable"""
        filename = f"Accounting_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Balances par entité
            for key, tb in accounting_data.items():
                if key.startswith('tb_'):
                    entity_name = key.replace('tb_', '').upper()
                    sheet_name = f'TB_{entity_name}'
                    tb.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._format_trial_balance_sheet(writer.book[sheet_name])
            
            # États financiers
            for key, statement in accounting_data.items():
                if key.startswith('bs_') or key.startswith('pl_'):
                    sheet_name = key.upper()
                    statement.to_excel(writer, sheet_name=sheet_name, index=False)
                    self._format_financial_statement_sheet(writer.book[sheet_name])
            
            # Méthodologie
            methodology = self._create_accounting_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport comptable généré: {filepath}")
        return str(filepath)
    
    def write_consolidation_report(self, consolidation_data: Dict[str, pd.DataFrame], 
                                 config: Dict) -> str:
        """Écrire le rapport de consolidation"""
        filename = f"Consolidation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Balance groupe
            if 'group_trial_balance' in consolidation_data:
                consolidation_data['group_trial_balance'].to_excel(
                    writer, sheet_name='Group_Trial_Balance', index=False
                )
                self._format_trial_balance_sheet(writer.book['Group_Trial_Balance'])
            
            # États consolidés
            if 'consolidated_balance_sheet' in consolidation_data:
                consolidation_data['consolidated_balance_sheet'].to_excel(
                    writer, sheet_name='Consolidated_BS', index=False
                )
                self._format_financial_statement_sheet(writer.book['Consolidated_BS'])
            
            if 'consolidated_income_statement' in consolidation_data:
                consolidation_data['consolidated_income_statement'].to_excel(
                    writer, sheet_name='Consolidated_PL', index=False
                )
                self._format_financial_statement_sheet(writer.book['Consolidated_PL'])
            
            # Liasse de consolidation
            if 'consolidation_package' in consolidation_data:
                consolidation_data['consolidation_package'].to_excel(
                    writer, sheet_name='Consolidation_Package', index=False
                )
                self._format_data_sheet(writer.book['Consolidation_Package'])
            
            # Détails FX
            if 'fx_translation_details' in consolidation_data:
                consolidation_data['fx_translation_details'].to_excel(
                    writer, sheet_name='FX_Translation', index=False
                )
                self._format_data_sheet(writer.book['FX_Translation'])
            
            # Intérêts minoritaires
            if 'minority_interests_details' in consolidation_data:
                consolidation_data['minority_interests_details'].to_excel(
                    writer, sheet_name='Minority_Interests', index=False
                )
                self._format_data_sheet(writer.book['Minority_Interests'])
            
            # Méthodologie
            methodology = self._create_consolidation_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport de consolidation généré: {filepath}")
        return str(filepath)
    
    def write_finrep_report(self, finrep_data: Dict[str, pd.DataFrame], 
                          config: Dict) -> str:
        """Écrire le rapport FINREP"""
        filename = f"FINREP_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # FINREP Balance Sheet
            if 'finrep_balance_sheet' in finrep_data:
                finrep_data['finrep_balance_sheet'].to_excel(
                    writer, sheet_name='FINREP_BS', index=False
                )
                self._format_finrep_sheet(writer.book['FINREP_BS'])
            
            # FINREP Profit & Loss
            if 'finrep_profit_loss' in finrep_data:
                finrep_data['finrep_profit_loss'].to_excel(
                    writer, sheet_name='FINREP_PL', index=False
                )
                self._format_finrep_sheet(writer.book['FINREP_PL'])
            
            # FINREP Equity Changes
            if 'finrep_equity_changes' in finrep_data:
                finrep_data['finrep_equity_changes'].to_excel(
                    writer, sheet_name='FINREP_Equity', index=False
                )
                self._format_data_sheet(writer.book['FINREP_Equity'])
            
            # FINREP Breakdown
            if 'finrep_breakdown' in finrep_data:
                finrep_data['finrep_breakdown'].to_excel(
                    writer, sheet_name='FINREP_Breakdown', index=False
                )
                self._format_data_sheet(writer.book['FINREP_Breakdown'])
            
            # Méthodologie
            methodology = self._create_finrep_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport FINREP généré: {filepath}")
        return str(filepath)
    
    def write_corep_report(self, corep_data: Dict[str, pd.DataFrame], 
                         config: Dict) -> str:
        """Écrire le rapport COREP"""
        filename = f"COREP_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # COREP Credit Risk
            if 'corep_credit_risk' in corep_data:
                corep_data['corep_credit_risk'].to_excel(
                    writer, sheet_name='COREP_Credit_Risk', index=False
                )
                self._format_corep_sheet(writer.book['COREP_Credit_Risk'])
            
            # COREP Own Funds
            if 'corep_own_funds' in corep_data:
                corep_data['corep_own_funds'].to_excel(
                    writer, sheet_name='COREP_Own_Funds', index=False
                )
                self._format_corep_sheet(writer.book['COREP_Own_Funds'])
            
            # COREP Capital Ratios
            if 'corep_capital_ratios' in corep_data:
                corep_data['corep_capital_ratios'].to_excel(
                    writer, sheet_name='COREP_Capital_Ratios', index=False
                )
                self._format_corep_sheet(writer.book['COREP_Capital_Ratios'])
            
            # COREP Leverage Ratio
            if 'corep_leverage_ratio' in corep_data:
                corep_data['corep_leverage_ratio'].to_excel(
                    writer, sheet_name='COREP_Leverage', index=False
                )
                self._format_corep_sheet(writer.book['COREP_Leverage'])
            
            # RWA Detail
            if 'rwa_detail' in corep_data:
                rwa_sample = corep_data['rwa_detail'].head(1000)
                rwa_sample.to_excel(writer, sheet_name='RWA_Detail', index=False)
                self._format_data_sheet(writer.book['RWA_Detail'])
            
            # Méthodologie
            methodology = self._create_corep_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport COREP généré: {filepath}")
        return str(filepath)
    
    def write_liquidity_report(self, liquidity_data: Dict[str, pd.DataFrame], 
                             config: Dict) -> str:
        """Écrire le rapport de liquidité"""
        filename = f"Liquidity_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # LCR
            if 'lcr_summary' in liquidity_data:
                liquidity_data['lcr_summary'].to_excel(
                    writer, sheet_name='LCR_Summary', index=False
                )
                self._format_liquidity_sheet(writer.book['LCR_Summary'])
            
            if 'lcr_detail' in liquidity_data:
                liquidity_data['lcr_detail'].to_excel(
                    writer, sheet_name='LCR_Detail', index=False
                )
                self._format_data_sheet(writer.book['LCR_Detail'])
            
            # NSFR
            if 'nsfr_summary' in liquidity_data:
                liquidity_data['nsfr_summary'].to_excel(
                    writer, sheet_name='NSFR_Summary', index=False
                )
                self._format_liquidity_sheet(writer.book['NSFR_Summary'])
            
            if 'nsfr_detail' in liquidity_data:
                liquidity_data['nsfr_detail'].to_excel(
                    writer, sheet_name='NSFR_Detail', index=False
                )
                self._format_data_sheet(writer.book['NSFR_Detail'])
            
            # ALMM
            if 'almm_maturity_ladder' in liquidity_data:
                liquidity_data['almm_maturity_ladder'].to_excel(
                    writer, sheet_name='ALMM_Maturity_Ladder', index=False
                )
                self._format_data_sheet(writer.book['ALMM_Maturity_Ladder'])
            
            if 'almm_summary' in liquidity_data:
                liquidity_data['almm_summary'].to_excel(
                    writer, sheet_name='ALMM_Summary', index=False
                )
                self._format_data_sheet(writer.book['ALMM_Summary'])
            
            # Méthodologie
            methodology = self._create_liquidity_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport de liquidité généré: {filepath}")
        return str(filepath)
    
    def write_ruba_report(self, ruba_data: Dict[str, pd.DataFrame], 
                        config: Dict) -> str:
        """Écrire le rapport RUBA"""
        filename = f"RUBA_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # RUBA Bilan
            if 'ruba_bilan' in ruba_data:
                ruba_data['ruba_bilan'].to_excel(
                    writer, sheet_name='RUBA_Bilan', index=False
                )
                self._format_ruba_sheet(writer.book['RUBA_Bilan'])
            
            # RUBA Compte de résultat
            if 'ruba_compte_resultat' in ruba_data:
                ruba_data['ruba_compte_resultat'].to_excel(
                    writer, sheet_name='RUBA_Compte_Resultat', index=False
                )
                self._format_ruba_sheet(writer.book['RUBA_Compte_Resultat'])
            
            # RUBA Ratios prudentiels
            if 'ruba_ratios_prudentiels' in ruba_data:
                ruba_data['ruba_ratios_prudentiels'].to_excel(
                    writer, sheet_name='RUBA_Ratios', index=False
                )
                self._format_ruba_sheet(writer.book['RUBA_Ratios'])
            
            # Méthodologie
            methodology = self._create_ruba_methodology(config)
            methodology.to_excel(writer, sheet_name='Methodology', index=False)
            self._format_methodology_sheet(writer.book['Methodology'])
        
        logger.info(f"Rapport RUBA généré: {filepath}")
        return str(filepath)
    
    def _format_summary_sheet(self, worksheet):
        """Formater une feuille de résumé"""
        # En-têtes
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Données
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_data_sheet(self, worksheet):
        """Formater une feuille de données"""
        # En-têtes
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Données
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        
        # Figer la première ligne
        worksheet.freeze_panes = 'A2'
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_trial_balance_sheet(self, worksheet):
        """Formater une feuille de balance"""
        self._format_data_sheet(worksheet)
        
        # Formatage spécial pour les montants
        for row in worksheet.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                if i >= 3:  # Colonnes de montants
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00_);(#,##0.00)'
                        if cell.value < 0:
                            cell.font = Font(color="FF0000")  # Rouge pour les négatifs
    
    def _format_financial_statement_sheet(self, worksheet):
        """Formater une feuille d'état financier"""
        self._format_data_sheet(worksheet)
        
        # Mise en évidence des totaux
        for row in worksheet.iter_rows(min_row=2):
            if any('TOTAL' in str(cell.value).upper() for cell in row):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    def _format_finrep_sheet(self, worksheet):
        """Formater une feuille FINREP"""
        self._format_data_sheet(worksheet)
        
        # Formatage spécial pour les codes FINREP
        for row in worksheet.iter_rows(min_row=2):
            first_cell = row[0]
            if first_cell.value and str(first_cell.value).isdigit():
                first_cell.font = Font(bold=True, color="000080")
    
    def _format_corep_sheet(self, worksheet):
        """Formater une feuille COREP"""
        self._format_data_sheet(worksheet)
        
        # Formatage spécial pour les codes COREP
        for row in worksheet.iter_rows(min_row=2):
            first_cell = row[0]
            if first_cell.value and 'C ' in str(first_cell.value):
                first_cell.font = Font(bold=True, color="008000")
    
    def _format_liquidity_sheet(self, worksheet):
        """Formater une feuille de liquidité"""
        self._format_data_sheet(worksheet)
        
        # Mise en évidence des ratios
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if 'ratio' in str(cell.value).lower():
                    cell.font = Font(bold=True)
    
    def _format_ruba_sheet(self, worksheet):
        """Formater une feuille RUBA"""
        self._format_data_sheet(worksheet)
        
        # Formatage spécial pour les codes RUBA
        for row in worksheet.iter_rows(min_row=2):
            first_cell = row[0]
            if first_cell.value and len(str(first_cell.value)) <= 3:
                first_cell.font = Font(bold=True, color="800080")
    
    def _format_methodology_sheet(self, worksheet):
        """Formater une feuille de méthodologie"""
        # En-têtes
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Données
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Ajuster la largeur des colonnes
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 50
    
    def _create_simulation_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie de simulation"""
        methodology_data = [
            {'Element': 'Objectif', 'Valeur': 'Simulation bancaire', 'Description': 'Génération de positions et flux réalistes'},
            {'Element': 'Horizon', 'Valeur': '2024-01-01 à 2024-12-31', 'Description': 'Période de simulation'},
            {'Element': 'Granularité', 'Valeur': 'Mensuelle', 'Description': 'Fréquence des calculs'},
            {'Element': 'Seed', 'Valeur': str(config.get('scenario_seed', 42)), 'Description': 'Graine de génération aléatoire'},
            {'Element': 'Entités', 'Valeur': 'EU, US, CN', 'Description': 'Filiales simulées'},
            {'Element': 'Produits', 'Valeur': 'Prêts, Dépôts, Titres, Dérivés', 'Description': 'Types de produits simulés'},
            {'Element': 'Paramètres_risque', 'Valeur': 'PD, LGD, CCF', 'Description': 'Paramètres de risque de crédit'},
            {'Element': 'Stages_ECL', 'Valeur': '1, 2, 3', 'Description': 'Stages de provisionnement IFRS 9'},
            {'Element': 'Devises', 'Valeur': 'EUR, USD, CNY', 'Description': 'Devises des entités'},
            {'Element': 'Limitations', 'Valeur': 'Simulation didactique', 'Description': 'Non destiné à un usage réel'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_accounting_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie comptable"""
        methodology_data = [
            {'Element': 'Normes', 'Valeur': 'IFRS simplifiées', 'Description': 'Normes comptables appliquées'},
            {'Element': 'Plan_comptable', 'Valeur': 'Simplifié bancaire', 'Description': 'Plan de comptes utilisé'},
            {'Element': 'Mapping_produits', 'Valeur': 'Automatique', 'Description': 'Allocation aux comptes par produit'},
            {'Element': 'Provisions', 'Valeur': 'ECL IFRS 9', 'Description': 'Méthode de provisionnement'},
            {'Element': 'Equilibrage', 'Valeur': 'Automatique', 'Description': 'Équilibrage des balances'},
            {'Element': 'Devise', 'Valeur': 'Locale par entité', 'Description': 'Devise de tenue des comptes'},
            {'Element': 'Etats_financiers', 'Valeur': 'Bilan et P&L', 'Description': 'États générés'},
            {'Element': 'Validation', 'Valeur': 'Contrôles automatiques', 'Description': 'Vérifications effectuées'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_consolidation_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie de consolidation"""
        methodology_data = [
            {'Element': 'Méthode', 'Valeur': 'Intégration globale', 'Description': 'Méthode de consolidation'},
            {'Element': 'Devise_base', 'Valeur': 'EUR', 'Description': 'Devise de consolidation'},
            {'Element': 'Conversion_bilan', 'Valeur': 'Taux de clôture', 'Description': 'Méthode de conversion des bilans'},
            {'Element': 'Conversion_PL', 'Valeur': 'Taux moyen', 'Description': 'Méthode de conversion des P&L'},
            {'Element': 'Eliminations', 'Valeur': 'Intercompagnies', 'Description': 'Éliminations effectuées'},
            {'Element': 'Interets_minoritaires', 'Valeur': 'Calculés', 'Description': 'Prise en compte des minoritaires'},
            {'Element': 'Pourcentages_detention', 'Valeur': 'EU:100%, US:80%, CN:60%', 'Description': 'Taux de détention'},
            {'Element': 'Ecarts_conversion', 'Valeur': 'En capitaux propres', 'Description': 'Traitement des écarts de change'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_finrep_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie FINREP"""
        methodology_data = [
            {'Element': 'Référentiel', 'Valeur': 'FINREP EBA', 'Description': 'Templates FINREP européens'},
            {'Element': 'Périmètre', 'Valeur': 'Groupe consolidé', 'Description': 'Périmètre de reporting'},
            {'Element': 'Normes_comptables', 'Valeur': 'IFRS', 'Description': 'Référentiel comptable'},
            {'Element': 'Mapping', 'Valeur': 'Comptes vers codes FINREP', 'Description': 'Correspondance utilisée'},
            {'Element': 'Etats_produits', 'Valeur': 'Bilan, P&L, Capitaux propres', 'Description': 'États financiers FINREP'},
            {'Element': 'Ventilations', 'Valeur': 'Par géographie et produit', 'Description': 'Analyses complémentaires'},
            {'Element': 'Fréquence', 'Valeur': 'Annuelle', 'Description': 'Périodicité de reporting'},
            {'Element': 'Devise', 'Valeur': 'EUR', 'Description': 'Devise de présentation'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_corep_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie COREP"""
        methodology_data = [
            {'Element': 'Référentiel', 'Valeur': 'COREP CRR3', 'Description': 'Templates COREP selon CRR3'},
            {'Element': 'Approche_retail', 'Valeur': 'IRB', 'Description': 'Méthode pour expositions retail'},
            {'Element': 'Approche_non_retail', 'Valeur': 'Standardisée', 'Description': 'Méthode pour expositions non-retail'},
            {'Element': 'Formules_IRB', 'Valeur': 'CRR3 Art. 153-154', 'Description': 'Formules réglementaires IRB'},
            {'Element': 'Pondérations_std', 'Valeur': 'CRR3 Art. 111-134', 'Description': 'Pondérations standardisées'},
            {'Element': 'Fonds_propres', 'Valeur': 'CET1, Tier1, Total', 'Description': 'Composants des fonds propres'},
            {'Element': 'Ratio_levier', 'Valeur': 'CRR3 Art. 429', 'Description': 'Calcul du ratio de levier'},
            {'Element': 'Plancher', 'Valeur': 'Non applicable', 'Description': 'Plancher de fonds propres'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_liquidity_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie de liquidité"""
        methodology_data = [
            {'Element': 'LCR_référentiel', 'Valeur': 'CRR3 Art. 412', 'Description': 'Ratio de couverture de liquidité'},
            {'Element': 'HQLA_classification', 'Valeur': 'Niveaux 1, 2A, 2B', 'Description': 'Classification des actifs liquides'},
            {'Element': 'Horizon_stress', 'Valeur': '30 jours', 'Description': 'Période de stress LCR'},
            {'Element': 'NSFR_référentiel', 'Valeur': 'CRR3 Art. 413', 'Description': 'Ratio de financement stable'},
            {'Element': 'ASF_RSF', 'Valeur': 'Facteurs réglementaires', 'Description': 'Facteurs de financement stable'},
            {'Element': 'ALMM_buckets', 'Valeur': 'Échéances standard', 'Description': 'Buckets de maturité ALMM'},
            {'Element': 'Taux_sortie', 'Valeur': 'Selon type de dépôt', 'Description': 'Taux de fuite des dépôts'},
            {'Element': 'Haircuts', 'Valeur': 'Selon qualité HQLA', 'Description': 'Décotes sur actifs liquides'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def _create_ruba_methodology(self, config: Dict) -> pd.DataFrame:
        """Créer la méthodologie RUBA"""
        methodology_data = [
            {'Element': 'Référentiel', 'Valeur': 'RUBA France', 'Description': 'Reporting uniforme bancaire français'},
            {'Element': 'Périmètre', 'Valeur': 'Groupe français', 'Description': 'Entités françaises du groupe'},
            {'Element': 'Structure', 'Valeur': 'Bilan et P&L simplifiés', 'Description': 'États RUBA produits'},
            {'Element': 'Codes_RUBA', 'Valeur': 'Nomenclature française', 'Description': 'Codification utilisée'},
            {'Element': 'Ratios_prudentiels', 'Valeur': 'Solvabilité, Levier, LCR, NSFR', 'Description': 'Ratios suivis'},
            {'Element': 'Fréquence', 'Valeur': 'Trimestrielle', 'Description': 'Périodicité de reporting'},
            {'Element': 'Autorité', 'Valeur': 'ACPR', 'Description': 'Autorité de supervision'},
            {'Element': 'Format', 'Valeur': 'Excel standardisé', 'Description': 'Format de transmission'}
        ]
        
        return pd.DataFrame(methodology_data)
    
    def add_charts_to_workbook(self, filepath: str, data_sheets: List[str]):
        """Ajouter des graphiques aux classeurs"""
        try:
            wb = openpyxl.load_workbook(filepath)
            
            for sheet_name in data_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Créer un graphique simple si des données numériques sont présentes
                    if ws.max_row > 1 and ws.max_column > 1:
                        chart = LineChart()
                        chart.title = f"Évolution - {sheet_name}"
                        chart.style = 13
                        chart.x_axis.title = 'Période'
                        chart.y_axis.title = 'Montant'
                        
                        # Données pour le graphique (première colonne numérique)
                        data = Reference(ws, min_col=2, min_row=1, max_row=min(ws.max_row, 20), max_col=2)
                        cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 20))
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        
                        # Positionner le graphique
                        ws.add_chart(chart, f"{chr(65 + ws.max_column + 1)}2")
            
            wb.save(filepath)
            logger.info(f"Graphiques ajoutés au fichier: {filepath}")
            
        except Exception as e:
            logger.warning(f"Impossible d'ajouter des graphiques: {e}")
    
    def create_master_summary(self, all_reports: Dict[str, str], config: Dict) -> str:
        """Créer un fichier de synthèse maître"""
        filename = f"Master_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Page de garde
            cover_data = [
                ['Banking Simulation & CRR3 Reporting', ''],
                ['', ''],
                ['Date de génération', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Période de référence', '2024-12-31'],
                ['Devise de base', 'EUR'],
                ['Périmètre', 'Groupe bancaire consolidé'],
                ['', ''],
                ['Rapports générés', ''],
                ['', '']
            ]
            
            for report_type, filepath_report in all_reports.items():
                cover_data.append([report_type.upper(), Path(filepath_report).name])
            
            cover_df = pd.DataFrame(cover_data, columns=['Élément', 'Valeur'])
            cover_df.to_excel(writer, sheet_name='Cover', index=False)
            
            # Index des rapports
            index_data = []
            for report_type, filepath_report in all_reports.items():
                index_data.append({
                    'Type de rapport': report_type.upper(),
                    'Nom du fichier': Path(filepath_report).name,
                    'Chemin complet': filepath_report,
                    'Taille (KB)': round(Path(filepath_report).stat().st_size / 1024, 2),
                    'Date de création': datetime.fromtimestamp(Path(filepath_report).stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            
            index_df = pd.DataFrame(index_data)
            index_df.to_excel(writer, sheet_name='Index', index=False)
            
            # Configuration utilisée
            config_data = []
            for key, value in config.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        config_data.append({
                            'Paramètre': f"{key}.{sub_key}",
                            'Valeur': str(sub_value)
                        })
                else:
                    config_data.append({
                        'Paramètre': key,
                        'Valeur': str(value)
                    })
            
            config_df = pd.DataFrame(config_data)
            config_df.to_excel(writer, sheet_name='Configuration', index=False)
            
            # Formater les feuilles
            self._format_data_sheet(writer.book['Cover'])
            self._format_data_sheet(writer.book['Index'])
            self._format_data_sheet(writer.book['Configuration'])
        
        logger.info(f"Fichier de synthèse maître créé: {filepath}")
        return str(filepath)
